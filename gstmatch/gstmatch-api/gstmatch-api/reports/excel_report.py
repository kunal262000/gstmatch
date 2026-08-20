"""
Generates a multi-sheet Excel report from a ReconciliationResult.
Supports both invoice-level and summary-level return reconciliations.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from models.schemas import ReconciliationResult, InvoiceCategory
from core.registry import get_recon_metadata

GREEN_FILL   = PatternFill("solid", fgColor="D1FAE5")
ORANGE_FILL  = PatternFill("solid", fgColor="FEF3C7")
RED_FILL     = PatternFill("solid", fgColor="FEE2E2")
HEADER_FILL  = PatternFill("solid", fgColor="1E293B")
ALT_FILL     = PatternFill("solid", fgColor="F8FAFC")

WHITE_FONT   = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT    = Font(bold=True)
TITLE_FONT   = Font(bold=True, size=14, color="1E293B")
DANGER_FONT  = Font(color="EF4444", bold=True)

THIN_BORDER  = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def _set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, headers: list, row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill      = HEADER_FILL
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
        cell.border    = THIN_BORDER


def _fmt_inr(amount: float) -> str:
    return f"₹{amount:,.0f}"


def _build_summary_sheet(ws, result: ReconciliationResult) -> None:
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    meta = get_recon_metadata(result.reconType or "gstr2b_pr")

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value     = f"GST Reconciliation — {result.period}"
    title_cell.font      = TITLE_FONT
    title_cell.alignment = CENTER
    title_cell.fill      = PatternFill("solid", fgColor="F0FDF4")

    ws.merge_cells("A2:D2")
    sub_cell = ws["A2"]
    sub_cell.value     = f"{result.businessName}  |  GSTIN: {result.gstin}  |  Type: {meta['name']}"
    sub_cell.alignment = CENTER
    sub_cell.font      = Font(color="475569", size=10)

    # Stats table
    stats = [
        ("",                   "Category",             "Count / Status",  ""),
        (GREEN_FILL,   "✅ Matched records",           result.summary.matched,          ""),
        (ORANGE_FILL,  "⚠️ Value / rate mismatch",     result.summary.mismatched,        ""),
        (RED_FILL,     f"❌ Missing in {meta['file2_label']}", result.summary.missingInGstr2b, ""),
        (ALT_FILL,     f"🔍 Missing in {meta['file1_label']}", result.summary.missingInPr,       ""),
        ("",           "Total records processed",      result.summary.totalInvoices,    ""),
        ("",           "Compliance score",             f"{result.summary.complianceScore}%", ""),
    ]

    for i, (fill, label, value, _) in enumerate(stats):
        row = i + 4
        if fill == "":
            ws.cell(row=row, column=2, value=label).font = BOLD_FONT
            ws.cell(row=row, column=3, value=value)
        elif label == "Category":
            _header_row(ws, ["", "Category", "Count / Status", ""], row)
        else:
            ws.cell(row=row, column=2, value=label).fill = fill
            ws.cell(row=row, column=3, value=value).fill = fill
            ws.cell(row=row, column=2).font = BOLD_FONT

    fin_diff = result.summary.financialDifference if result.summary.financialDifference is not None else (result.summary.totalItcAtRisk or 0.0)
    itc_row = len(stats) + 6
    ws.merge_cells(f"A{itc_row}:D{itc_row}")
    itc_cell = ws[f"A{itc_row}"]
    itc_cell.value     = f"{meta['financial_metric_label'].upper()}:  {_fmt_inr(fin_diff)}"
    itc_cell.fill      = RED_FILL if fin_diff > 0 else GREEN_FILL
    itc_cell.font      = Font(bold=True, size=13, color="EF4444" if fin_diff > 0 else "10B981")
    itc_cell.alignment = CENTER

    _set_col_widths(ws, [4, 38, 18, 4])


def _build_mismatch_sheet(ws, result: ReconciliationResult) -> None:
    ws.title = "Mismatches"
    ws.sheet_view.showGridLines = False
    meta = get_recon_metadata(result.reconType or "gstr2b_pr")

    if result.summarySections:
        headers = ["Section ID", "Section Name", "Description", f"{meta['file1_label']} (₹)", f"{meta['file2_label']} (₹)", "Difference (₹)", "Status"]
        _header_row(ws, headers)
        for i, sec in enumerate(result.summarySections):
            row = i + 2
            fill = RED_FILL if sec.status != "matched" else GREEN_FILL
            ws.cell(row=row, column=1, value=sec.sectionId).border = THIN_BORDER
            ws.cell(row=row, column=2, value=sec.sectionName).border = THIN_BORDER
            ws.cell(row=row, column=3, value=sec.description).border = THIN_BORDER
            ws.cell(row=row, column=4, value=sec.file1Value).border = THIN_BORDER
            ws.cell(row=row, column=5, value=sec.file2Value).border = THIN_BORDER
            ws.cell(row=row, column=6, value=sec.totalDifference).border = THIN_BORDER
            ws.cell(row=row, column=7, value=sec.status).fill = fill
            ws.cell(row=row, column=7).border = THIN_BORDER
        _set_col_widths(ws, [14, 30, 36, 18, 18, 16, 14])
        return

    headers = [meta["party_label"], "GSTIN", "Invoice No", "Date",
               f"{meta['file1_label']} Amount (₹)", f"{meta['file2_label']} Amount (₹)", "Difference (₹)", "Action Required"]
    _header_row(ws, headers)

    mismatched = [inv for inv in result.invoices if inv.category == InvoiceCategory.mismatched]
    for i, inv in enumerate(mismatched):
        row  = i + 2
        fill = ORANGE_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFBEB")
        diff = inv.difference or 0

        ws.cell(row=row, column=1, value=inv.supplierName).border = THIN_BORDER
        ws.cell(row=row, column=2, value=inv.gstin).border        = THIN_BORDER
        ws.cell(row=row, column=3, value=inv.invoiceNo).border    = THIN_BORDER
        ws.cell(row=row, column=4, value=inv.invoiceDate).border  = THIN_BORDER
        ws.cell(row=row, column=5, value=inv.yourAmount).border   = THIN_BORDER
        ws.cell(row=row, column=6, value=inv.gstr2bAmount or 0).border = THIN_BORDER
        ws.cell(row=row, column=7, value=diff).fill               = fill
        ws.cell(row=row, column=7).border                         = THIN_BORDER
        ws.cell(row=row, column=8, value="Verify invoice / rate difference").border = THIN_BORDER

    _set_col_widths(ws, [24, 18, 16, 13, 20, 20, 16, 26])


def _build_missing_invoices_sheet(ws, result: ReconciliationResult) -> None:
    ws.title = "Missing Invoices"
    ws.sheet_view.showGridLines = False
    meta = get_recon_metadata(result.reconType or "gstr2b_pr")

    missing_gstr2b = [inv for inv in result.invoices if inv.category == InvoiceCategory.missing_in_gstr2b]
    missing_pr = [inv for inv in result.invoices if inv.category == InvoiceCategory.missing_in_pr]

    if not missing_gstr2b and not missing_pr:
        ws.merge_cells("A1:H1")
        ws["A1"] = "No missing invoices found"
        ws["A1"].font = Font(color="64748b", size=12, italic=True)
        _set_col_widths(ws, [24, 18, 16, 13, 20, 20, 16, 26])
        return

    all_missing = missing_gstr2b + missing_pr
    headers = ["Supplier", "GSTIN", "Invoice No", "Date",
               f"{meta['file1_label']} Amount (₹)", f"{meta['file2_label']} Amount (₹)",
               "Tax Component (₹)", "Category"]
    _header_row(ws, headers)

    all_filled = 0
    for i, inv in enumerate(all_missing):
        row = i + 2
        fill = ORANGE_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFBEB")
        tax1 = float(inv.igst or 0) + float(inv.cgst or 0) + float(inv.sgst or 0)

        ws.cell(row=row, column=1, value=inv.supplierName).border = THIN_BORDER
        ws.cell(row=row, column=2, value=inv.gstin).border = THIN_BORDER
        ws.cell(row=row, column=3, value=inv.invoiceNo).border = THIN_BORDER
        ws.cell(row=row, column=4, value=inv.invoiceDate).border = THIN_BORDER
        ws.cell(row=row, column=5, value=inv.yourAmount).border = THIN_BORDER
        ws.cell(row=row, column=6, value=inv.gstr2bAmount or 0).border = THIN_BORDER
        ws.cell(row=row, column=7, value=round(tax1, 2)).border = THIN_BORDER
        cat_label = "Missing in GSTR-2B" if inv.category == InvoiceCategory.missing_in_gstr2b else "Missing in Purchase Register"
        ws.cell(row=row, column=8, value=cat_label).border = THIN_BORDER
        all_filled += 1

    _set_col_widths(ws, [24, 18, 16, 13, 20, 20, 16, 26])


def generate_excel_report(result: ReconciliationResult) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    _build_summary_sheet(ws_summary, result)

    ws_mismatch = wb.create_sheet("Mismatches")
    _build_mismatch_sheet(ws_mismatch, result)

    ws_missing = wb.create_sheet("Missing Invoices")
    _build_missing_invoices_sheet(ws_missing, result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
