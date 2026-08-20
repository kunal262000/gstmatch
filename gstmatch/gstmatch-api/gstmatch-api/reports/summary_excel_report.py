"""
NEW FILE — place at: gstmatch-api/gstmatch-api/reports/summary_excel_report.py

Excel report for SUMMARY-level reconciliations (GSTR-3B vs GSTR-1,
GSTR-1 vs GSTR-3B, GSTR-9 vs Books, GSTR-9C vs Books).

Reuses the exact colour palette from reports/excel_report.py for visual
consistency across the product.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.schemas import SummaryReconciliationResult

GREEN_FILL  = PatternFill("solid", fgColor="D1FAE5")
RED_FILL    = PatternFill("solid", fgColor="FEE2E2")
HEADER_FILL = PatternFill("solid", fgColor="1E293B")
ALT_FILL    = PatternFill("solid", fgColor="F8FAFC")
WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT   = Font(bold=True)
DANGER_FONT = Font(color="EF4444", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"), right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"), bottom=Side(style="thin", color="E2E8F0"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def _fmt(amount: float) -> str:
    return f"₹{amount:,.0f}"


def generate_summary_excel(result: SummaryReconciliationResult) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = f"{result.file1Label} vs {result.file2Label} — {result.period}"
    title.font = Font(bold=True, size=14, color="1E293B")
    title.alignment = CENTER
    title.fill = PatternFill("solid", fgColor="F0FDF4")

    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = f"{result.businessName}  |  GSTIN: {result.gstin}"
    sub.alignment = CENTER
    sub.font = Font(color="475569", size=10)

    stats_row = 4
    stats = [
        (f"Total in {result.file1Label}", _fmt(result.totalFile1Value)),
        (f"Total in {result.file2Label}", _fmt(result.totalFile2Value)),
        ("Net Difference", _fmt(result.totalDifference)),
        ("Matched Sections", str(result.matchedSections)),
        ("Mismatched Sections", str(result.mismatchedSections)),
    ]
    for i, (label, val) in enumerate(stats):
        ws.cell(row=stats_row + i, column=2, value=label).font = BOLD_FONT
        cell = ws.cell(row=stats_row + i, column=3, value=val)
        if "Difference" in label and result.totalDifference != 0:
            cell.font = DANGER_FONT

    header_row = stats_row + len(stats) + 2
    headers = ["Section", f"{result.file1Label} (₹)", f"{result.file2Label} (₹)",
               "Difference (₹)", "IGST Diff", "CGST Diff", "SGST Diff", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill, cell.font, cell.alignment, cell.border = HEADER_FILL, WHITE_FONT, CENTER, THIN_BORDER

    for i, li in enumerate(result.lineItems):
        row = header_row + 1 + i
        fill = RED_FILL if li.status == "mismatch" else (GREEN_FILL if i % 2 == 0 else ALT_FILL)
        values = [li.section, li.file1Value, li.file2Value, li.difference,
                 li.igstDiff, li.cgstDiff, li.sgstDiff,
                 "⚠ Mismatch" if li.status == "mismatch" else "✓ Matched"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill, cell.border = fill, THIN_BORDER
            if col == 4 and li.difference != 0:
                cell.font = DANGER_FONT

    for i, w in enumerate([30, 18, 18, 16, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
